# -*- coding:utf-8 -*-
import sys
reload(sys)
sys.setdefaultencoding('utf-8')

import openpyxl


wb = openpyxl.load_workbook("test.xlsx")


for sheetName in wb.get_sheet_names():
    if sheetName == u'表1':
        sheet = wb[sheetName]
        break

p = {}
p.setdefault(r"用电周期",[])
e = {}
e.setdefault(r"用电量",[])
for row_num in range(2,sheet.max_row):
    period_num = sheet.cell(row=row_num,column=9).value
    p["用电周期"].append(period_num)
    e_sum_num = sheet.cell(row=row_num,column=12).value
    e["用电量"].append(e_sum_num)

print p
print e

resultwb = openpyxl.Workbook()
resultsheet = resultwb.get_active_sheet()
resultsheet.title = "result"


#Todo: open a text file and write the contents of data